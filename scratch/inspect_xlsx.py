import openpyxl
import pandas as pd
import os

file_path = "C:/Users/GCMS/Documents/GitHub/lims/Report_ISTD_PARAMETER_NEW.xlsx"

print("File size:", os.path.getsize(file_path))

# Check sheet names using openpyxl
wb = openpyxl.load_workbook(file_path, read_only=True)
print("Sheets in workbook:", wb.sheetnames)

# Let's inspect the first sheet
sheet_name = wb.sheetnames[0]
df = pd.read_excel(file_path, sheet_name=sheet_name)
print(f"\n--- First sheet '{sheet_name}' head ---")
print(df.head(20))
print(df.info())
